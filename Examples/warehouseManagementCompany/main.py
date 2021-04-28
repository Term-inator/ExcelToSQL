import openpyxl
from excelToSQL import Table

employer = Table("employer", 0)

warehouse = Table("wareHouse", 0)

_employer = openpyxl.load_workbook("employer.xlsx", read_only=True)
_warehouse = openpyxl.load_workbook("warehouse.xlsx", read_only=True)

_employer_sheet_names = _employer.sheetnames
sheet = _employer[_employer_sheet_names[0]]

employer.link([("empID", None, False),
               ("empNo", None, False),
               ("wareHouseNo", None, False),
               ("empName", None, False),
               ("empSex", None, False),
               ("salary", None, False)],
              sheet, [1, 2, 3, 4, 5, 6], (2, sheet.max_row))

employer.unique([])
employer.generate_sql()

_warehouse_sheet_names = _warehouse.sheetnames
sheet = _warehouse[_warehouse_sheet_names[0]]

warehouse.link([("wareHouseID", None, False),
                ("wareHouseNo", None, False),
                ("city", None, False),
                ("area", None, False),
                ("buildTime", None, False)],
               sheet, [1, 2, 3, 4, 5], (2, sheet.max_row))

warehouse.unique([])
warehouse.generate_sql()

sql = open("sql.txt", "a+", encoding="utf-8")

sql.write(employer.get_sql())
sql.write(warehouse.get_sql())

sql.close()

