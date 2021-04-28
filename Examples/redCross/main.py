import openpyxl
from excelToSQL import Table

office = Table("office", 0)
office.add_id()

item = Table("item", 0)
item.add_id()

donation = Table("donation", 2)

book_send = openpyxl.load_workbook("发放物资（2.13-2.14）.xlsx", read_only=True)
book_donate = openpyxl.load_workbook("捐赠物资（2.13-2.14）.xlsx", read_only=True)


sheet_send_names = book_send.sheetnames
sheet = book_send[sheet_send_names[0]]

office.link([("name", None, False)],
            sheet, [1], (4, sheet.max_row))

item.link([("name", None, False),
           ("specification", None, False),
           ("unit", None, False)],
          sheet, [2, 3, 4], (4, sheet.max_row))

donation.link([("office_id", 0, True),
               ("item_id", 0, True),
               ("donate", 0, False),
               ("amount", None, False),
               ("time", "2.13", False)],
              sheet, [0, 0, 0, 5, 0], (4, sheet.max_row))


sheet = book_send[sheet_send_names[1]]

office.link([("name", None, False)],
            sheet, [1], (4, sheet.max_row))
item.link([("name", None, False),
           ("specification", None, False),
           ("unit", None, False)],
          sheet, [2, 3, 4], (4, sheet.max_row))

donation.link([("office_id", 0, True),
               ("item_id", 0, True),
               ("donate", 0, False),
               ("amount", None, False),
               ("time", "2.14", False)],
              sheet, [0, 0, 0, 5, 0], (4, sheet.max_row))


sheet_donate_names = book_donate.sheetnames
sheet = book_donate[sheet_donate_names[0]]

office.link([("name", None, False)], sheet,
            [1], (5, sheet.max_row))

item.link([("name", None, False),
           ("specification", None, False),
           ("unit", None, False)], sheet,
          [2, 3, 4], (5, sheet.max_row))

donation.link([("office_id", 0, True),
               ("item_id", 0, True),
               ("donate", 1, False),
               ("amount", None, False),
               ("time", "2.13", False)],
              sheet, [0, 0, 0, 5, 0], (5, sheet.max_row))


sheet = book_donate[sheet_donate_names[1]]

office.link([("name", None, False)],
            sheet, [1], (4, sheet.max_row))

item.link([("name", None, False),
           ("specification", None, False),
           ("unit", None, False)],
          sheet, [2, 3, 4], (4, sheet.max_row))

donation.link([("office_id", 0, True),
               ("item_id", 0, True),
               ("donate", 1, False),
               ("amount", None, False),
               ("time", "2.14", False)],
              sheet, [0, 0, 0, 5, 0], (4, sheet.max_row))

office.unique(["name"])
item.unique(["name", "specification"])
donation.unique([])


######################################
sheet = book_send[sheet_send_names[0]]
donation.reference(("office_id", "id"), office, "name", sheet, 1, (4, sheet.max_row))
donation.reference(("item_id", "id"), item, "name", sheet, 2, (4, sheet.max_row))

sheet = book_send[sheet_send_names[1]]
donation.reference(("office_id", "id"), office, "name", sheet, 1, (4, sheet.max_row))
donation.reference(("item_id", "id"), item, "name", sheet, 2, (4, sheet.max_row))

sheet = book_donate[sheet_donate_names[0]]
donation.reference(("office_id", "id"), office, "name", sheet, 1, (5, sheet.max_row))
donation.reference(("item_id", "id"), item, "name", sheet, 2, (5, sheet.max_row))

sheet = book_donate[sheet_donate_names[1]]
donation.reference(("office_id", "id"), office, "name", sheet, 1, (4, sheet.max_row))
donation.reference(("item_id", "id"), item, "name", sheet, 2, (4, sheet.max_row))

office.generate_sql()
item.generate_sql()
donation.generate_sql()

sql = open("sql.txt", "a+", encoding="utf-8")

sql.write(office.get_sql())
sql.write(item.get_sql())
sql.write(donation.get_sql())

sql.close()
